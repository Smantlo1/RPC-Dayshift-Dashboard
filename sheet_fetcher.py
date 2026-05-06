"""
sheet_fetcher.py — Fetch the live xlsx from SharePoint, parse with openpyxl, cache.

Auth strategy (tried in order, first success wins):
  1. Stored browser cookies (FedAuth / rtFa from settings.json)
       — most reliable for ADFS / Azure AD tenants like Walmart
  2. PowerShell Invoke-WebRequest -UseDefaultCredentials
       — native WinHTTP stack, sometimes works on federated intranet
  3. requests + Windows Negotiate SSPI
       — works only on classic on-prem SharePoint w/ NTLM, kept as last resort
  4. Fail with clear auth-failure flag so the UI shows the cookie setup form

Cookie setup: one-time. User copies FedAuth + rtFa from browser DevTools,
pastes into the dashboard form. Stored in settings.json alongside the DB.
"""

import asyncio
import io
import json
import logging
import os
import subprocess
import tempfile
import time
from datetime import date, datetime
from pathlib import Path
from typing import Optional

from routine_data import TRACKING_SHEET_RAW_URL, TRACKING_SHEET_VIEW_URL

log = logging.getLogger(__name__)

CACHE_TTL     = 600   # seconds (10 min)
MAX_ROWS      = 500
MAX_COLS      = 40
SETTINGS_FILE = Path(__file__).parent / "settings.json"

# ── Status cell colouring ────────────────────────────────────────────────────
_STATUS_MAP: list[tuple[frozenset, str]] = [
    (frozenset({"on track", "complete", "done", "ordered", "yes"}),   "cell-ok"),
    (frozenset({"overdue", "past due", "blocked", "critical", "no"}), "cell-bad"),
    (frozenset({"at risk", "in progress", "pending", "tbd"}),         "cell-warn"),
]


def status_class(val: str) -> str:
    key = val.strip().lower()
    for keywords, cls in _STATUS_MAP:
        if key in keywords:
            return cls
    return ""


# ── Settings (cookie persistence) ───────────────────────────────────────────
def _load_settings() -> dict:
    try:
        return json.loads(SETTINGS_FILE.read_text(encoding="utf-8"))
    except (FileNotFoundError, json.JSONDecodeError, OSError):
        return {}


def _save_settings(data: dict) -> None:
    SETTINGS_FILE.write_text(json.dumps(data, indent=2), encoding="utf-8")


def load_sp_cookie() -> Optional[str]:
    """Return stored SharePoint cookie string, or None."""
    return _load_settings().get("sp_cookie") or None


def save_sp_cookie(raw: str) -> None:
    """Persist a cookie string to settings.json and bust the cache."""
    s = _load_settings()
    s["sp_cookie"] = raw.strip()
    _save_settings(s)
    _cache.update({"sheets": None, "error": None, "fetched_at": 0.0, "auth_failure": False})


def clear_sp_cookie() -> None:
    """Remove stored cookie and bust the cache."""
    s = _load_settings()
    s.pop("sp_cookie", None)
    _save_settings(s)
    _cache.update({"sheets": None, "error": None, "fetched_at": 0.0, "auth_failure": False})


# ── In-memory cache ──────────────────────────────────────────────────────────
_cache: dict = {
    "sheets":      None,
    "error":       None,
    "fetched_at":  0.0,
    "auth_failure": False,
}


def _is_fresh() -> bool:
    return (
        _cache["sheets"] is not None
        and (time.monotonic() - _cache["fetched_at"]) < CACHE_TTL
    )


# ── Cell formatter ───────────────────────────────────────────────────────────
def _fmt(val) -> str:
    if val is None:
        return ""
    if isinstance(val, bool):
        return "Yes" if val else "No"
    if isinstance(val, (datetime, date)):
        return val.strftime("%m/%d/%Y")
    if isinstance(val, float):
        return str(int(val)) if val == int(val) else f"{val:.2f}"
    return str(val).strip()


# ── Per-strategy fetch helpers ───────────────────────────────────────────────

def _valid_xlsx(data: bytes) -> bool:
    return len(data) >= 4 and data[:4] == b"PK\x03\x04"


def _strategy_cookie(cookie: str) -> tuple[Optional[bytes], Optional[str]]:
    """Use stored browser session cookies (FedAuth / rtFa)."""
    try:
        import requests
    except ImportError:
        return None, "requests not installed"

    try:
        r = requests.get(
            TRACKING_SHEET_RAW_URL,
            headers={
                "Cookie": cookie,
                "User-Agent": (
                    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                    "AppleWebKit/537.36 RPC-Dashboard/1.0"
                ),
                "Accept": (
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet,"
                    "application/octet-stream,*/*"
                ),
            },
            timeout=25,
            allow_redirects=True,
        )
    except Exception as exc:
        return None, str(exc)

    if r.status_code in (401, 403):
        return None, f"HTTP {r.status_code} — cookie may have expired"
    if r.status_code != 200:
        return None, f"HTTP {r.status_code}"

    ct = r.headers.get("content-type", "")
    if "html" in ct.lower() or not _valid_xlsx(r.content):
        return None, "Got HTML login page — cookie has expired"

    return r.content, None


def _strategy_powershell() -> tuple[Optional[bytes], Optional[str]]:
    """Use PowerShell's WinHTTP stack (works on some federated intranets)."""
    tmp = tempfile.mktemp(suffix=".xlsx")
    try:
        result = subprocess.run(
            [
                "powershell", "-NonInteractive", "-Command",
                (
                    f'Invoke-WebRequest -Uri "{TRACKING_SHEET_RAW_URL}" '
                    f'-UseDefaultCredentials -OutFile "{tmp}" '
                    f'-TimeoutSec 20 -ErrorAction Stop'
                ),
            ],
            capture_output=True,
            text=True,
            timeout=28,
            encoding="utf-8",
            errors="replace",
        )
        if result.returncode != 0:
            msg = (result.stderr or result.stdout or "non-zero exit").strip()
            return None, f"PowerShell: {msg[:200]}"

        with open(tmp, "rb") as fh:
            data = fh.read()

        if _valid_xlsx(data):
            return data, None
        return None, "PowerShell download returned non-xlsx content (auth redirect?)"

    except subprocess.TimeoutExpired:
        return None, "PowerShell timed out after 28s"
    except FileNotFoundError:
        return None, "powershell not found"
    except Exception as exc:
        return None, str(exc)
    finally:
        try:
            os.unlink(tmp)
        except OSError:
            pass


def _strategy_sspi() -> tuple[Optional[bytes], Optional[str]]:
    """requests + Windows Negotiate SSPI (classic on-prem SharePoint only)."""
    try:
        import requests
    except ImportError:
        return None, "requests not installed"

    session = requests.Session()
    session.headers.update({
        "User-Agent": "RPC-Dashboard/1.0",
        "Accept": (
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet,"
            "application/octet-stream,*/*"
        ),
    })

    auth = None
    try:
        from requests_negotiate_sspi import HttpNegotiateAuth
        auth = HttpNegotiateAuth()
    except ImportError:
        pass

    try:
        r = session.get(TRACKING_SHEET_RAW_URL, auth=auth, timeout=25, allow_redirects=True)
    except requests.exceptions.ConnectionError:
        return None, "Cannot reach SharePoint — check VPN / Eagle WiFi"
    except requests.exceptions.Timeout:
        return None, "Request timed out (25s)"
    except Exception as exc:
        return None, str(exc)

    if r.status_code in (401, 403):
        return None, f"HTTP {r.status_code}"
    if r.status_code != 200:
        return None, f"HTTP {r.status_code}"
    ct = r.headers.get("content-type", "")
    if "html" in ct.lower() or not _valid_xlsx(r.content):
        return None, "Got HTML response (login page)"
    return r.content, None


# ── Main fetch orchestrator ───────────────────────────────────────────────────

def _fetch_bytes() -> tuple[Optional[bytes], Optional[str], bool]:
    """
    Returns (xlsx_bytes, error_msg, auth_failure).
    auth_failure=True means: show the cookie setup form.
    """
    # ── Strategy 1: stored browser cookie ────────────────────────────────
    cookie = load_sp_cookie()
    if cookie:
        log.debug("Trying strategy 1: stored browser cookie")
        data, err = _strategy_cookie(cookie)
        if data:
            log.info("Sheet fetched via stored cookie")
            return data, None, False
        log.debug("Cookie strategy failed: %s", err)
        # Cookie exists but bounced — likely expired, flag as auth failure
        return None, (
            f"Your SharePoint session cookie has expired ({err}). "
            "Please refresh it below."
        ), True

    # ── Strategy 2: PowerShell UseDefaultCredentials ─────────────────────
    log.debug("Trying strategy 2: PowerShell UseDefaultCredentials")
    data, err = _strategy_powershell()
    if data:
        log.info("Sheet fetched via PowerShell")
        return data, None, False
    log.debug("PowerShell strategy failed: %s", err)

    # ── Strategy 3: requests + SSPI ──────────────────────────────────────
    log.debug("Trying strategy 3: requests + SSPI")
    data, err = _strategy_sspi()
    if data:
        log.info("Sheet fetched via SSPI")
        return data, None, False
    log.debug("SSPI strategy failed: %s", err)

    # ── All strategies failed → prompt for cookie ─────────────────────────
    return None, (
        "Automatic authentication failed. "
        "Paste your SharePoint session cookie below to connect."
    ), True


# ── xlsx parser ───────────────────────────────────────────────────────────────

def _parse(xlsx_bytes: bytes) -> tuple[Optional[list], Optional[str]]:
    try:
        import openpyxl
    except ImportError:
        return None, "openpyxl is not installed"

    try:
        wb = openpyxl.load_workbook(
            io.BytesIO(xlsx_bytes), read_only=True, data_only=True
        )
    except Exception as exc:
        return None, f"Could not open workbook: {exc}"

    sheets = []
    for name in wb.sheetnames:
        ws = wb[name]
        all_rows: list = []
        for row in ws.iter_rows(values_only=True):
            all_rows.append(row)
            if len(all_rows) > MAX_ROWS + 20:
                break

        if not all_rows:
            continue

        header_idx = next(
            (i for i, r in enumerate(all_rows) if any(c is not None for c in r)), 0
        )
        raw_headers = list(all_rows[header_idx])[:MAX_COLS]

        last_col = max(
            (i for i, h in enumerate(raw_headers) if h is not None and str(h).strip()),
            default=0,
        ) + 1
        raw_headers = raw_headers[:last_col]
        headers = [_fmt(h) or f"Col {i+1}" for i, h in enumerate(raw_headers)]

        data_rows: list[list[str]] = []
        truncated = False
        for row in all_rows[header_idx + 1:]:
            cells = [_fmt(c) for c in list(row)[:last_col]]
            if any(cells):
                data_rows.append(cells)
            if len(data_rows) >= MAX_ROWS:
                truncated = True
                break

        sheets.append({
            "name":      name,
            "headers":   headers,
            "rows":      data_rows,
            "row_count": len(data_rows),
            "truncated": truncated,
        })

    wb.close()
    return (sheets, None) if sheets else (None, "Workbook has no data in any sheet.")


# ── Cache runner ──────────────────────────────────────────────────────────────

def _run_fetch() -> None:
    raw, err, auth_fail = _fetch_bytes()
    if err:
        _cache.update({
            "sheets": None, "error": err,
            "fetched_at": time.monotonic(), "auth_failure": auth_fail,
        })
        return
    sheets, err = _parse(raw)
    _cache.update({
        "sheets": sheets, "error": err,
        "fetched_at": time.monotonic(), "auth_failure": False,
    })


# ── Public async API ──────────────────────────────────────────────────────────

async def get_sheet_data(force: bool = False) -> dict:
    """
    Returns dict: sheets, error, fetched_at, auth_failure, has_cookie, view_url
    """
    if not force and _is_fresh():
        return {**_cache, "has_cookie": bool(load_sp_cookie()), "view_url": TRACKING_SHEET_VIEW_URL}

    loop = asyncio.get_event_loop()
    await loop.run_in_executor(None, _run_fetch)
    return {**_cache, "has_cookie": bool(load_sp_cookie()), "view_url": TRACKING_SHEET_VIEW_URL}
